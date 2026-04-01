import streamlit as st
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
import io
from openpyxl import load_workbook
import os
import json

# --- 1. CẤU HÌNH & DANH SÁCH NĂM ---
st.set_page_config(page_title="Hồ Sơ Học Tập Số", page_icon="🎓", layout="wide")

YEAR_LIST = [f"{y}-{y+1}" for y in range(2025, 2030)]

def init_firebase():
    if not firebase_admin._apps:
        try:
            # ƯU TIÊN 1: Chạy trên Server Render (Lấy từ biến môi trường)
            if "FIREBASE_JSON" in os.environ:
                key_dict = json.loads(os.environ["FIREBASE_JSON"])
                if "private_key" in key_dict:
                    key_dict["private_key"] = key_dict["private_key"].replace("\\n", "\n")
            # ƯU TIÊN 2: Chạy trên máy tính cá nhân (Lấy từ file secrets)
            else:
                key_dict = dict(st.secrets["firebase"])
                key_dict["private_key"] = key_dict["private_key"].replace("\\n", "\n")
                
            cred = credentials.Certificate(key_dict)
            firebase_admin.initialize_app(cred)
        except Exception as e:
            st.error(f"Lỗi kết nối Firebase: {e}")
            st.stop()
    return firestore.client()

# --- 2. CSS GIAO DIỆN (DÀNH CHO TOÀN TRANG & ADMIN) ---
st.markdown("""
<style>
    /* Ẩn râu ria */
    #MainMenu, header, footer, .stAppDeployButton {display: none !important;}
    [data-testid="stSidebar"] {display: none;}
    .block-container {padding: 0.5rem 0.5rem 2rem 0.5rem !important;}
    
    /* Header Admin */
    .main-header {
        background: blue;
        padding: 15px; border-radius: 12px; color: white; 
        text-align: center; font-weight: 700; font-size: 16px;
        box-shadow: 0 4px 10px rgba(0,0,0,0.2); margin-bottom: 15px;
        text-transform: uppercase; letter-spacing: 1px;
    }
    
    /* Report Card */
    .report-card {
        background: white; padding: 15px; border: 1px solid #ddd;
        border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.08); 
        margin-bottom: 15px; color: #333; position: relative;
    }
    .year-tag {
        position: absolute; top: 10px; right: 10px;
        background: #e3f2fd; color: #1565c0; padding: 4px 8px;
        border-radius: 6px; font-size: 12px; font-weight: bold;
    }
    
    /* Grid Tổng kết */
    .summary-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; margin-top: 15px; }
    .summary-item { background: #f8f9fa; padding: 10px; border-radius: 8px; border-left: 4px solid #2c5364; text-align: center; }
    .summary-val { font-size: 15px; font-weight: bold; color: #333; margin-top: 2px; display:block;}
    
    /* Table & Button */
    .stTable { font-size: 13px; }
    div[data-testid="stTable"] td { padding: 8px 2px !important; }
    .stButton>button { width: 100%; border-radius: 10px; height: 48px; font-weight: bold; }
    
    /* Admin Zone */
    .admin-zone { border: 1px dashed #ccc; padding: 15px; border-radius: 10px; background: #fdfdfd; margin-top: 20px;}
    .config-box { background: #e8f5e9; padding: 10px; border-radius: 8px; border: 1px solid #c8e6c9; margin-bottom: 15px; text-align: center;}
    .copyright {
        background: #33CCFF;
        text-align: center; font-size: 13px; padding: 5px;
    }
</style>
""", unsafe_allow_html=True)

# --- 3. HÀM XỬ LÝ DATABASE & CẤU HÌNH ---
def safe_str(val):
    if pd.isna(val) or str(val).lower() in ['nan', 'none', '']: return ""
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def load_excel_robust(file):
    try: return pd.read_excel(file, sheet_name=None)
    except:
        try: file.seek(0); dfs = pd.read_html(file); return {f"Sheet {i+1}": df for i, df in enumerate(dfs)}
        except: return None

def get_current_year_config(db):
    try:
        doc = db.collection('system_config').document('settings').get()
        if doc.exists: return doc.to_dict().get('default_year', '2024-2025')
    except: pass
    return '2024-2025'

def set_current_year_config(db, year):
    db.collection('system_config').document('settings').set({'default_year': year}, merge=True)

def get_activation_fee(db):
    try:
        doc = db.collection('system_config').document('settings').get()
        if doc.exists: return doc.to_dict().get('activation_fee', 15000)
    except: pass
    return 15000

def set_activation_fee(db, fee):
    db.collection('system_config').document('settings').set({'activation_fee': fee}, merge=True)

def delete_data_year(db, collection, year, cls, sem=None):
    cnt = 0
    try:
        ref = db.collection(collection)
        query = ref.where('year', '==', year)
        if cls != "Tất cả": query = query.where('cls', '==', cls)
        if sem: query = query.where('sem', '==', sem)
        
        batch = db.batch(); b_cnt = 0
        for doc in query.stream():
            batch.delete(doc.reference)
            b_cnt += 1; cnt += 1
            if b_cnt >= 400: batch.commit(); batch = db.batch(); b_cnt = 0
        if b_cnt > 0: batch.commit()
    except Exception as e: st.error(f"Lỗi xóa: {e}")
    return cnt

def upload_firebase(db, file, year, sem, cls, type_file):
    count = 0
    try:
        batch = db.batch(); b_cnt = 0
        
        if type_file == 'score':
            data = load_excel_robust(file)
            if not data: return 0
            for sname, df in data.items():
                if any(x in str(sname).lower() for x in ["hướng dẫn", "bìa"]): continue
                h_idx = -1
                for i, row in df.iterrows():
                    if row.astype(str).str.contains("Mã học sinh", case=False).any(): h_idx = i; break
                if h_idx != -1:
                    df.columns = df.iloc[h_idx]; df = df.iloc[h_idx+1:]
                    cols = df.columns.tolist()
                    idx_ma = next((i for i,c in enumerate(cols) if "Mã học sinh" in str(c)), -1)
                    if idx_ma != -1:
                        for _, row in df.iterrows():
                            ma = safe_str(row.iloc[idx_ma])
                            if len(ma) > 3:
                                try:
                                    ten = safe_str(row.iloc[idx_ma-2])
                                    doc_st_id = f"{ma}_{year}"
                                    ref_st = db.collection('students').document(doc_st_id)
                                    snap = ref_st.get()
                                    st_data = {'id': ma, 'name': ten, 'cls': cls, 'year': year}
                                    if not snap.exists: st_data['active'] = 0
                                    batch.set(ref_st, st_data, merge=True)
                                except: pass

                                def g(o): 
                                    try: return safe_str(row.iloc[idx_ma+o])
                                    except: return ""
                                sub = str(sname).strip().replace("/", "-")
                                doc_id = f"{ma}_{year}_{sem}_{sub}"
                                batch.set(db.collection('scores').document(doc_id), {
                                    'id': ma, 'year': year, 'sem': sem, 'cls': cls, 'sub': sub,
                                    'tx': "  ".join([g(k) for k in range(1,10) if g(k)]),
                                    'gk': g(16), 'ck': g(26), 'tb': g(27), 
                                    'cn': (g(28) if sem=='HK2' else "")
                                })
                                count += 1; b_cnt += 1
                                if b_cnt >= 300: batch.commit(); batch = db.batch(); b_cnt = 0
            batch.commit()
        elif type_file == 'summary':
            try: df = pd.read_excel(file)
            except: df = pd.read_csv(file)
            if 'Mã học sinh' not in df.columns:
                for i, r in df.iterrows():
                    if r.astype(str).str.contains("Mã học sinh").any(): df.columns = df.iloc[i]; df = df.iloc[i+1:]; break
            df.columns = df.columns.str.strip()
            has_loai = 'Loại TK' in df.columns
            for _, row in df.iterrows():
                ma = safe_str(row.get('Mã học sinh'))
                if len(ma) > 3:
                    cur_sem = sem
                    if has_loai:
                        v = safe_str(row.get('Loại TK')).upper()
                        if '1' in v: cur_sem = 'HK1'
                        elif '2' in v: cur_sem = 'HK2'
                        elif 'CN' in v or 'NAM' in v: cur_sem = 'CN'
                    doc_id = f"{ma}_{year}_{cur_sem}_sum"
                    batch.set(db.collection('summary').document(doc_id), {
                        'id': ma, 'year': year, 'sem': cur_sem, 'cls': cls,
                        'ht': safe_str(row.get('Học tập')), 'rl': safe_str(row.get('Rèn luyện')),
                        'v': safe_str(row.get('Vắng')), 'dh': safe_str(row.get('Danh hiệu')),
                        'kq': safe_str(row.get('Kết quả'))
                    })
                    count += 1; b_cnt += 1
                    if b_cnt >= 300: batch.commit(); batch = db.batch(); b_cnt = 0
            batch.commit()
    except Exception as e: st.error(f"Lỗi: {e}")
    return count

# --- 4. ADMIN UI ---
def view_admin(db):
    st.markdown('<div class="main-header">🛠️ QUẢN TRỊ VIÊN</div>', unsafe_allow_html=True)
    if st.button("Đăng xuất"): st.session_state.page = 'login'; st.rerun()
    
    if st.text_input("Mật khẩu:", type="password") == "admin123":
        current_db_year = get_current_year_config(db)
        current_fee = get_activation_fee(db)
        fee_formatted = f"{current_fee:,}".replace(',', '.')
        
        st.markdown(f"""<div class="config-box"><b>Năm học đang kích hoạt: {current_db_year} | Phí kích hoạt: {fee_formatted} VNĐ</b></div>""", unsafe_allow_html=True)
        
        col_y, col_f = st.columns(2)
        with col_y:
            year_sel = st.selectbox("📅 Năm làm việc:", YEAR_LIST, index=YEAR_LIST.index(current_db_year) if current_db_year in YEAR_LIST else 0)
            if st.button("📌 Đặt làm Mặc định", use_container_width=True):
                set_current_year_config(db, year_sel)
                st.success(f"Đã đặt {year_sel} làm mặc định!"); st.rerun()
        with col_f:
            fee_input = st.number_input("💰 Cấu hình Phí kích hoạt (VNĐ):", min_value=0, value=int(current_fee), step=1000)
            if st.button("📌 Cập nhật Phí", use_container_width=True):
                set_activation_fee(db, int(fee_input))
                st.success(f"Đã cập nhật phí thành {f'{int(fee_input):,}'.replace(',', '.')} VNĐ!"); st.rerun()

        st.markdown("---")
        t1, t2, t3 = st.tabs(["UPLOADER", "KÍCH HOẠT", "XÓA DỮ LIỆU"])
        
        with t1:
            st.caption(f"Upload vào năm: **{year_sel}**")
            cls = st.selectbox("Lớp:", [f"Lớp {i}" for i in range(6, 13)])
            c1, c2 = st.columns(2)
            f1 = c1.file_uploader(f"Điểm HK1 {cls}", key="f1")
            f2 = c1.file_uploader(f"Điểm HK2 {cls}", key="f2")
            tk = st.file_uploader(f"Tổng Kết {cls}", key="tk")
            
            if st.button("LƯU DỮ LIỆU", type="primary"):
                with st.spinner(f"Đang xử lý {year_sel}..."):
                    c = 0
                    if f1: c += upload_firebase(db, f1, year_sel, "HK1", cls, 'score')
                    if f2: c += upload_firebase(db, f2, year_sel, "HK2", cls, 'score')
                    if tk: c += upload_firebase(db, tk, year_sel, "HK1", cls, 'summary')
                    st.success(f"Đã lưu {c} bản ghi.")

        with t2:
            flt = st.selectbox("Lọc Lớp:", ["Tất cả"] + [f"Lớp {i}" for i in range(6, 13)])
            ref = db.collection('students').where('year', '==', year_sel)
            if flt != "Tất cả": ref = ref.where('cls', '==', flt)
            
            docs = list(ref.stream())
            data = [{"id_doc": d.id, **d.to_dict()} for d in docs]
            
            if data:
                df = pd.DataFrame(data)
                if 'active' not in df.columns: df['active'] = 0
                df['active'] = df['active'].apply(lambda x: bool(x))
                
                df = df.sort_values(by=['cls', 'name'])
                df.insert(0, 'STT', range(1, len(df) + 1))
                
                edited = st.data_editor(df[['active', 'STT', 'id', 'name', 'cls']], 
                                      column_config={
                                          "active": st.column_config.CheckboxColumn("Kích hoạt", default=False),
                                          "STT": st.column_config.NumberColumn("STT", width="small", disabled=True),
                                          "id": st.column_config.TextColumn("Mã HS", disabled=True),
                                          "name": st.column_config.TextColumn("Họ tên", disabled=True),
                                          "cls": st.column_config.TextColumn("Lớp", disabled=True)
                                      },
                                      hide_index=True, use_container_width=True)
                
                if st.button("LƯU TRẠNG THÁI"):
                    batch = db.batch(); b_cnt = 0
                    for i, r in edited.iterrows():
                        doc_key = f"{r['id']}_{year_sel}"
                        batch.update(db.collection('students').document(doc_key), {'active': 1 if r['active'] else 0})
                        b_cnt += 1
                        if b_cnt >= 300: batch.commit(); batch = db.batch(); b_cnt = 0
                    batch.commit()
                    st.success("Đã lưu!")
            else: st.warning(f"Chưa có dữ liệu năm {year_sel}.")

        with t3:
            st.warning(f"Đang xóa dữ liệu năm: {year_sel}")
            del_cls = st.selectbox("Lớp xóa:", ["Tất cả"] + [f"Lớp {i}" for i in range(6, 13)], key="del")
            c1, c2 = st.columns(2)
            
            with c1:
                st.markdown("**1. Xóa Điểm Chi Tiết:**")
                d_hk1 = st.checkbox("Xóa Điểm HK1")
                d_hk2 = st.checkbox("Xóa Điểm HK2")
                
            with c2:
                st.markdown("**2. Xóa Tổng Kết:**")
                d_thk1 = st.checkbox("Xóa TK HK1")
                d_thk2 = st.checkbox("Xóa TK HK2")       
                d_tcn  = st.checkbox("Xóa TK Cả Năm")    
            
            st.markdown("**3. Khác:**")
            d_all = st.checkbox("Xóa Tài khoản HS (Reset năm)")
            
            if st.button("🚨 THỰC HIỆN XÓA", type="primary"):
                with st.spinner("Deleting..."):
                    if d_hk1: delete_data_year(db, 'scores', year_sel, del_cls, 'HK1')
                    if d_hk2: delete_data_year(db, 'scores', year_sel, del_cls, 'HK2')
                    
                    if d_thk1: delete_data_year(db, 'summary', year_sel, del_cls, 'HK1')
                    if d_thk2: delete_data_year(db, 'summary', year_sel, del_cls, 'HK2') 
                    if d_tcn:  delete_data_year(db, 'summary', year_sel, del_cls, 'CN')  
                    
                    if d_all: delete_data_year(db, 'students', year_sel, del_cls)
                    st.success("Đã xóa xong!")

        # --- MODULE GHÉP MÃ HS (CHỈ ADMIN THẤY) ---
        st.markdown("---")
        st.markdown('<div class="admin-zone">', unsafe_allow_html=True)
        st.markdown("<h3 style='text-align: center; color: #ff4500; text-shadow: 0 0 10px #ff0000;'>⚡ CÔNG CỤ TỰ ĐỘNG GHÉP MÃ HỌC SINH</h3>", unsafe_allow_html=True)
        st.info("Khu vực ẩn dành riêng cho Admin: Upload file chứa mã HS và file điểm n môn để hệ thống tự động dò tìm 'Họ và tên' và điền 'Mã học sinh' vào tất cả các sheet. LƯU Ý: Để giữ nguyên 100% định dạng gốc, File Điểm bắt buộc phải là đuôi .xlsx")
        
        col_diem, col_ma = st.columns(2)
        f_diem = col_diem.file_uploader("1. File Điểm (Bắt buộc .xlsx)", type=['xlsx'], key="f_diem_ghep")
        f_ma = col_ma.file_uploader("2. File Mã Học Sinh", type=['xlsx', 'xls', 'csv'], key="f_ma_ghep")
        
        if f_diem and f_ma:
            if st.button("🚀 BÙNG NỔ HỢP NHẤT DỮ LIỆU", type="primary", use_container_width=True):
                with st.spinner("Đang quét và hợp nhất dữ liệu trên tất cả các sheet... 🔥"):
                    try:
                        def safe_norm(val):
                            if pd.isna(val) or val is None: return ""
                            return str(val).strip().lower()

                        try: df_ma = pd.read_excel(f_ma)
                        except: df_ma = pd.read_csv(f_ma)
                        
                        df_ma['Họ và tên'] = df_ma['Họ và tên'].apply(safe_norm)
                        ma_dict = dict(zip(df_ma['Họ và tên'], df_ma['Mã học sinh']))
                        
                        file_diem_bytes = io.BytesIO(f_diem.getvalue())
                        
                        try:
                            wb = load_workbook(file_diem_bytes)
                        except Exception as zip_err:
                            st.error("❌ Lỗi định dạng! File điểm của thầy có thể đang bị lỗi hoặc là file .xls cũ bị đổi đuôi. Xin hãy mở file điểm bằng Excel, chọn 'Save As' và lưu lại chuẩn với định dạng 'Excel Workbook (*.xlsx)' rồi upload lại nhé!")
                            st.stop()
                        
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            
                            header_row_idx = -1
                            col_name_idx = -1
                            col_id_idx = -1
                            
                            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
                                for col_idx, cell in enumerate(row, start=1):
                                    val = safe_norm(cell.value)
                                    if "họ và tên" in val or "họ tên" in val:
                                        header_row_idx = row_idx
                                        col_name_idx = col_idx
                                    if "mã học sinh" in val:
                                        col_id_idx = col_idx
                                        
                                if header_row_idx != -1 and col_id_idx != -1:
                                    break 
                            
                            if header_row_idx != -1 and col_name_idx != -1 and col_id_idx != -1:
                                for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                                    name_cell = ws.cell(row=row_idx, column=col_name_idx)
                                    id_cell = ws.cell(row=row_idx, column=col_id_idx)
                                    
                                    temp_name = safe_norm(name_cell.value)
                                    
                                    if temp_name and temp_name in ma_dict:
                                        id_cell.value = ma_dict[temp_name]
                                            
                        output = io.BytesIO()
                        wb.save(output)
                        processed_data = output.getvalue()
                        
                        st.success("Hợp nhất thành công! Tải file kết quả ngay bên dưới 🚀")
                        st.download_button(
                            label="📥 TẢI FILE KẾT QUẢ VỀ MÁY",
                            data=processed_data,
                            file_name="Diem_Da_Ghep_Ma_Giu_Nguyen_Dinh_Dang.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"Có lỗi xảy ra trong quá trình xử lý: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

# --- 5. HỌC SINH UI ---
def view_student(db):
    st.markdown("""
    <style>
        .block-container { max-width: 650px !important; padding-top: 2rem !important; }
        .stApp { background: linear-gradient(135deg, #0a0000 0%, #2a0000 50%, #000000 100%) !important; background-attachment: fixed !important; }
        .stMarkdown p, .stMarkdown strong { color: #ffffff !important; text-shadow: 0 0 5px rgba(255, 69, 0, 0.5) !important; }
        .neon-title { text-align: center; font-size: 24px; font-weight: 900; color: #fff; text-transform: uppercase; letter-spacing: 2px; text-shadow: 0 0 10px #ff0000, 0 0 20px #ff4500, 0 0 40px #ff0000; margin-bottom: 25px; padding-bottom: 15px; border-bottom: 2px solid rgba(255, 69, 0, 0.4); line-height: 1.5; }
        .stTextInput label, .stSelectbox label { color: #ff4500 !important; font-weight: bold !important; font-size: 14px !important; text-transform: uppercase !important; text-shadow: 0 0 5px rgba(255, 0, 0, 0.5) !important; }
        .stRadio div[role="radiogroup"] label p { color: #ffeb3b !important; font-weight: 900 !important; font-size: 16px !important; text-shadow: 0 0 8px #ff0000 !important; }
        .stTextInput input, div[data-baseweb="select"] > div { background-color: rgba(0, 0, 0, 0.7) !important; color: #ffeb3b !important; border: 2px solid #ff4500 !important; border-radius: 8px !important; box-shadow: 0 0 15px rgba(255, 69, 0, 0.4) inset, 0 0 10px rgba(255, 0, 0, 0.4) !important; font-size: 16px !important; font-weight: bold !important; }
        .stTextInput input { text-align: center !important; }
        div[data-baseweb="select"] * { color: #ffeb3b !important; font-weight: bold; }
        .stButton > button { background: linear-gradient(45deg, #990000, #ff4500) !important; color: white !important; border: 2px solid #ff0000 !important; border-radius: 8px !important; font-weight: 900 !important; font-size: 18px !important; text-transform: uppercase !important; letter-spacing: 2px !important; box-shadow: 0 0 20px rgba(255, 0, 0, 0.6) !important; transition: all 0.3s ease !important; height: 50px !important; }
        .stButton > button:hover { transform: scale(1.05) !important; box-shadow: 0 0 30px rgba(255, 69, 0, 1), 0 0 10px #fff inset !important; background: linear-gradient(45deg, #ff0000, #ff7300) !important; border-color: #ffeb3b !important; color: #fff !important; }
        .stTable { background-color: rgba(20,0,0,0.8) !important; border-radius: 10px; overflow: hidden; border: 1px solid #ff4500 !important; box-shadow: 0 0 15px rgba(255,0,0,0.4) !important; }
        .stTable th { background-color: #550000 !important; color: #ffeb3b !important; border-bottom: 2px solid #ff4500 !important; text-align: center !important; text-shadow: 0 0 5px #ff0000 !important; }
        .stTable td { border-bottom: 1px solid rgba(255, 69, 0, 0.3) !important; text-align: center !important; color: #ffffff !important; font-weight: bold !important; text-shadow: 0 0 5px #ff4500 !important; }
        .stAlert { background-color: rgba(20,0,0,0.9) !important; color: #ffeb3b !important; border: 1px solid #ff4500 !important; box-shadow: 0 0 10px rgba(255,0,0,0.5) !important;}
        .report-card { background: rgba(20,0,0,0.8) !important; border: 1px solid #ff4500 !important; box-shadow: 0 0 15px rgba(255,0,0,0.4) !important; }
        .report-card div { color: #fff !important; text-shadow: 0 0 5px rgba(255,255,255,0.5); }
        .year-tag { background: #550000 !important; color: #ffeb3b !important; border: 1px solid #ff4500 !important; }
        .summary-item { background: rgba(20,0,0,0.8) !important; border: 1px solid #ff4500 !important; border-left: 4px solid #ff0000 !important; box-shadow: 0 0 10px rgba(255,0,0,0.3) !important; }
        .summary-val { color: #ffeb3b !important; text-shadow: 0 0 5px #ffeb3b; }
        .copyright { background: transparent !important; color: #ff4500 !important; text-shadow: 0 0 5px #ff0000; border-top: 1px dashed #ff4500; margin-top: 40px;}
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="neon-title">🔥 TRA CỨU ĐIỂM SỐ 🔥<br><span style="font-size: 16px; color: #ffeb3b; text-shadow: none;">Trường PT DTNT THCS&THPT Tuy Đức</span></div>', unsafe_allow_html=True)
    
    default_year = get_current_year_config(db)
    current_fee = get_activation_fee(db) 
    fee_k = current_fee // 1000 

    if 'show_activation' not in st.session_state:
        st.session_state.show_activation = False

    if 'user' not in st.session_state:
        try: idx = YEAR_LIST.index(default_year)
        except: idx = 0
        year_login = st.selectbox("Năm học:", YEAR_LIST, index=idx)
        mid = st.text_input("Mã Học Sinh:", placeholder="VD: 2411...").strip()
        
        if st.button("TRA CỨU", type="primary", use_container_width=True):
            if not mid:
                st.error("Vui lòng nhập Mã Học Sinh!")
            else:
                doc_key = f"{mid}_{year_login}"
                doc = db.collection('students').document(doc_key).get()
                if not doc.exists: 
                    st.error(f"❌ Không tìm thấy dữ liệu năm {year_login}! Liên hệ admin zalo: 0383477162 để đăng kí tài khoản. (Cân nhắc phí kích hoạt {fee_k}k/năm học)")
                    st.session_state.show_activation = False
                elif doc.to_dict().get('active') != 1: 
                    st.warning(f"🔒 Tài khoản của bạn chưa được kích hoạt cho năm {year_login}.")
                    st.session_state.show_activation = True
                    st.session_state.temp_mid = mid
                else:
                    st.session_state.user = doc.to_dict()
                    st.session_state.year_view = year_login
                    st.session_state.show_activation = False
                    st.rerun()

        if st.session_state.get('show_activation'):
            temp_mid = st.session_state.get('temp_mid', '')
            fee_formatted = f"{current_fee:,}".replace(',', '.')
            
            st.markdown("""
            <div style="background-color: rgba(20,0,0,0.8); border: 2px solid #ff4500; border-radius: 12px; padding: 20px; text-align: center; margin-top: 15px; box-shadow: 0 0 20px rgba(255,0,0,0.4);">
                <h3 style="color: #ffeb3b; text-shadow: 0 0 10px #ff0000; margin-bottom: 10px; text-transform: uppercase;">🚀 Hướng dẫn kích hoạt tài khoản</h3>
                <p style="color: #fff; font-size: 15px; margin-bottom: 20px;">Sử dụng App Ngân hàng quét mã QR bên dưới.<br><i style="color: #4CAF50;">(Hệ thống đã tự động điền Mã Học Sinh vào nội dung chuyển khoản)</i></p>
            """, unsafe_allow_html=True)
            
            qr_url = f"https://img.vietqr.io/image/agribank-5300215042850-compact2.png?amount={current_fee}&addInfo={temp_mid}&accountName=LUONG%20VAN%20GIOI"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.image(qr_url, use_container_width=True)
                
            st.markdown(f"""
                <hr style="border-color: rgba(255,69,0,0.5); margin: 20px 0;">
                <p style="color: #fff; font-size: 16px; font-weight: bold; text-transform: uppercase;">Hoặc chuyển khoản thủ công:</p>
                <div style="background: rgba(0,0,0,0.5); padding: 15px; border-radius: 8px; border: 1px dashed #ffeb3b; text-align: left; display: inline-block;">
                    <p style="margin: 8px 0; color: #fff; font-size: 15px;">🏦 Ngân hàng: <b style="color: #4CAF50;">Agribank</b></p>
                    <p style="margin: 8px 0; color: #fff; font-size: 15px;">💳 Số tài khoản: <b style="color: #ffeb3b; font-size: 18px;">5300215042850</b></p>
                    <p style="margin: 8px 0; color: #fff; font-size: 15px;">👤 Chủ tài khoản: <b style="color: #2196F3;">LUONG VAN GIOI</b></p>
                    <p style="margin: 8px 0; color: #fff; font-size: 15px;">💰 Số tiền: <b style="color: #ff9800;">{fee_formatted} VNĐ</b></p>
                    <p style="margin: 8px 0; color: #fff; font-size: 15px;">📝 Nội dung CK: <b style="color: #ff4500; font-size: 18px; background: rgba(255,69,0,0.2); padding: 2px 8px; border-radius: 4px;">{temp_mid}</b></p>
                </div>
                <p style="color: #ff9999; margin-top: 15px; font-size: 13px;"><i>* Sau khi CK thành công, vui lòng chụp màn hình gửi Zalo <b>0383477162</b> để Admin duyệt!</i></p>
            </div>
            """, unsafe_allow_html=True)
            
    else:
        u = st.session_state.user
        year_view = st.session_state.year_view
        
        st.markdown(f"""
        <div class="report-card">
            <span class="year-tag">{year_view}</span>
            <div style="text-align:center; font-weight:bold; color:#ffeb3b !important; font-size:18px;">
                {u.get('name')}
            </div>
            <div style="text-align:center; font-size:14px; margin-top: 5px;">
                Mã: {u.get('id')} | Lớp: {u.get('cls')}
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        ky = st.radio("", ["Học kỳ 1", "Học kỳ 2 & Cả năm"], horizontal=True)
        sem = "HK1" if "1" in ky else "HK2"
        
        docs = db.collection('scores').where('id', '==', u['id']).where('year', '==', year_view).where('sem', '==', sem).stream()
        data = [d.to_dict() for d in docs]
        
        if data:
            df = pd.DataFrame(data)
            
            def sort_priority(row):
                s = str(row['sub']).lower()
                if 'toán' in s: return 0
                if 'văn' in s or 'ngữ văn' in s: return 1
                if 'anh' in s or 'ngoại ngữ' in s: return 2
                
                eval_subs = ['thể chất', 'gdtc', 'quốc phòng', 'gdqp', 'trải nghiệm', 'hđtn', 'địa phương', 'nghệ thuật', 'âm nhạc', 'mỹ thuật']
                if any(x in s for x in eval_subs): return 20 
                return 10 
            
            df['priority'] = df.apply(sort_priority, axis=1)
            df = df.sort_values(by=['priority', 'sub'])
            
            df['STT'] = range(1, len(df)+1)
            
            rn = {'sub': 'Môn', 'tx': 'TX', 'gk': 'GK', 'ck': 'CK', 'tb': 'TB', 'cn': 'CN'}
            cols = ['STT', 'Môn', 'TX', 'GK', 'CK', 'TB']
            if sem == 'HK2': cols.append('CN')
            st.table(df.rename(columns=rn)[cols].set_index('STT'))
        else: st.info("Chưa có điểm.")
        
        doc_tk = f"{u['id']}_{year_view}_{sem}_sum"
        tk = db.collection('summary').document(doc_tk).get()
        tk_d = tk.to_dict() if tk.exists else {}
        
        def card(l, v): return f'<div class="summary-item"><small style="color: #ff9999">{l}</small><div class="summary-val">{v if v else "-"}</div></div>'
        st.markdown(f"**TỔNG KẾT {sem}**")
        if tk_d: st.markdown(f"""<div class="summary-grid">{card('Học lực', tk_d.get('ht'))}{card('Hạnh kiểm', tk_d.get('rl'))}{card('Vắng', tk_d.get('v'))}{card('Danh hiệu', tk_d.get('dh'))}</div>""", unsafe_allow_html=True)
        
        if sem == 'HK2':
            doc_cn = f"{u['id']}_{year_view}_CN_sum"
            cn = db.collection('summary').document(doc_cn).get()
            cn_d = cn.to_dict() if cn.exists else {}
            if cn_d:
                st.markdown("---")
                st.markdown("**CẢ NĂM**")
                st.markdown(f"""<div class="summary-grid">{card('Học lực', cn_d.get('ht'))}{card('Hạnh kiểm', cn_d.get('rl'))}{card('Danh hiệu', cn_d.get('dh'))}<div class="summary-item" style="border-color:#ff0000; background:rgba(255,0,0,0.2)"><small style="color:#ff4500">KẾT QUẢ</small><div class="summary-val" style="color:#ff0000; text-shadow: 0 0 10px #ff0000;">{cn_d.get('kq')}</div></div></div>""", unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        if c1.button("🔙 Đổi Năm"): del st.session_state.user; st.rerun()
        if c2.button("Thoát"): del st.session_state.user; st.rerun()

    st.markdown('<div class="admin-zone" style="text-align:center; border:none; margin-top:50px; background: transparent;">', unsafe_allow_html=True)
    if st.button("⚙️", key="adm_btn"): st.session_state.page = 'admin'; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    if 'page' not in st.session_state: st.session_state.page = 'login'
    try:
        db = init_firebase()
        if st.session_state.page == 'admin': view_admin(db)
        else: view_student(db)
    except Exception as e: st.error("Lỗi hệ thống."); print(e)
    
    st.markdown('<div class="copyright">Copyright©2026 - Lương Văn Giỏi - 0383477162</div>', unsafe_allow_html=True)
